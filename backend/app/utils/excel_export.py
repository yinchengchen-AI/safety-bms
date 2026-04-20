from io import BytesIO
from typing import Any
from urllib.parse import quote

from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Font


def export_excel_response(filename: str, headers: list[str], rows: list[list[Any]]) -> Response:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Sheet1"

    # Header
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Rows
    for row in rows:
        ws.append(row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_length:
                    max_length = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    encoded_filename = quote(filename)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )
