from fastapi import APIRouter, Depends, File, Query, UploadFile
from sqlalchemy.orm import Session

from app.core.constants import CustomerStatus, PermissionCode
from app.core.exceptions import BusinessError, NotFoundError, PermissionDeniedError
from app.crud.customer import crud_customer
from app.db.session import get_db
from app.dependencies import require_permissions
from app.models.customer import Customer
from app.models.user import User
from app.schemas.common import PageResponse, ResponseMsg
from app.schemas.customer import (
    CustomerContactCreate,
    CustomerContactOut,
    CustomerCreate,
    CustomerFollowUpCreate,
    CustomerFollowUpOut,
    CustomerListOut,
    CustomerOut,
    CustomerUpdate,
)
from app.utils.data_scope import apply_data_scope, check_data_scope
from app.utils.excel_export import export_excel_response
from app.utils.export_mappings import CUSTOMER_STATUS_MAP, map_value
from app.utils.pagination import make_page_response

router = APIRouter(prefix="/customers", tags=["客户管理"])


@router.get("", response_model=PageResponse[CustomerListOut])
def list_customers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    status: CustomerStatus | None = None,
    keyword: str | None = None,
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_READ)),
    db: Session = Depends(get_db),
):
    skip = (page - 1) * page_size
    query = db.query(Customer).filter(Customer.is_deleted == False)
    if status:
        query = query.filter(Customer.status == status)
    if keyword:
        query = query.filter(Customer.name.ilike(f"%{keyword}%"))
    query = apply_data_scope(query, Customer, current_user)
    total = query.count()
    items = query.order_by(Customer.created_at.desc()).offset(skip).limit(page_size).all()
    return make_page_response(total, items, page, page_size)


@router.get("/export")
def export_customers(
    status: CustomerStatus | None = None,
    keyword: str | None = None,
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_READ)),
    db: Session = Depends(get_db),
):
    query = db.query(Customer).filter(Customer.is_deleted == False)
    if status:
        query = query.filter(Customer.status == status)
    if keyword:
        query = query.filter(Customer.name.ilike(f"%{keyword}%"))
    query = apply_data_scope(query, Customer, current_user)
    items = query.order_by(Customer.created_at.desc()).all()
    headers = [
        "客户名称",
        "统一信用代码",
        "行业",
        "规模",
        "地址",
        "联系人",
        "联系电话",
        "状态",
        "创建时间",
    ]
    rows = []
    for c in items:
        rows.append(
            [
                c.name,
                c.credit_code or "",
                c.industry or "",
                c.scale or "",
                c.address or "",
                c.contact_name or "",
                c.contact_phone or "",
                map_value(c.status.value if c.status else "", CUSTOMER_STATUS_MAP),
                c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
            ]
        )
    from datetime import UTC, datetime

    return export_excel_response(
        f"customers_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.xlsx", headers, rows
    )


@router.post("", response_model=CustomerOut, status_code=201)
def create_customer(
    body: CustomerCreate,
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_CREATE)),
    db: Session = Depends(get_db),
):
    return crud_customer.create(db, obj_in=body, created_by=current_user.id)


@router.get("/{customer_id}", response_model=CustomerOut)
def get_customer(
    customer_id: int,
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_READ)),
    db: Session = Depends(get_db),
):
    customer = crud_customer.get(db, id=customer_id)
    if not customer or customer.is_deleted:
        raise NotFoundError("客户")
    if not check_data_scope(customer, current_user):
        raise BusinessError("无权查看该记录", status_code=403)
    return customer


@router.patch("/{customer_id}", response_model=CustomerOut)
def update_customer(
    customer_id: int,
    body: CustomerUpdate,
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_UPDATE)),
    db: Session = Depends(get_db),
):
    customer = crud_customer.get(db, id=customer_id)
    if not customer or customer.is_deleted:
        raise NotFoundError("客户")
    if not check_data_scope(customer, current_user):
        raise PermissionDeniedError()
    return crud_customer.update(db, db_obj=customer, obj_in=body)


@router.delete("/{customer_id}", response_model=ResponseMsg)
def delete_customer(
    customer_id: int,
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_DELETE)),
    db: Session = Depends(get_db),
):
    customer = crud_customer.get(db, id=customer_id)
    if not customer or customer.is_deleted:
        raise NotFoundError("客户")
    if not check_data_scope(customer, current_user):
        raise PermissionDeniedError()
    crud_customer.soft_delete(db, customer_id=customer_id)
    return {"message": "删除成功"}


@router.post("/import", response_model=dict)
def import_customers(
    file: UploadFile = File(...),
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_CREATE)),
    db: Session = Depends(get_db),
):
    import io

    from openpyxl import load_workbook

    content = file.file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
    except Exception as exc:
        raise BusinessError(f"无法解析Excel文件: {exc}") from exc

    ws = wb.active
    if not ws:
        raise BusinessError("Excel文件为空")

    headers = [cell.value for cell in ws[1]]
    expected = ["客户名称", "统一信用代码", "行业", "规模", "地址", "联系人", "联系电话"]
    if headers[: len(expected)] != expected:
        raise BusinessError(f"Excel表头不符合要求，期望: {expected}")

    success = 0
    errors: list[dict] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = str(row[0]).strip() if row[0] else None
        if not name:
            continue
        credit_code = str(row[1]).strip() if row[1] else None
        industry = str(row[2]).strip() if row[2] else None
        scale = str(row[3]).strip() if row[3] else None
        address = str(row[4]).strip() if row[4] else None
        contact_name = str(row[5]).strip() if row[5] else None
        contact_phone = str(row[6]).strip() if row[6] else None
        try:
            existing = (
                db.query(Customer)
                .filter(Customer.name == name, Customer.is_deleted == False)
                .first()
            )
            if existing:
                errors.append({"row": idx, "error": f"客户名称 '{name}' 已存在"})
                continue
            crud_customer.create(
                db,
                obj_in=CustomerCreate(
                    name=name,
                    credit_code=credit_code,
                    industry=industry,
                    scale=scale,
                    address=address,
                    contact_name=contact_name,
                    contact_phone=contact_phone,
                ),
                created_by=current_user.id,
            )
            success += 1
        except Exception as exc:
            errors.append({"row": idx, "error": str(exc)})

    return {"success": success, "failed": len(errors), "errors": errors}


@router.post("/{customer_id}/contacts", response_model=CustomerContactOut, status_code=201)
def add_contact(
    customer_id: int,
    body: CustomerContactCreate,
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_UPDATE)),
    db: Session = Depends(get_db),
):
    customer = crud_customer.get(db, id=customer_id)
    if not customer or customer.is_deleted:
        raise NotFoundError("客户")
    if not check_data_scope(customer, current_user):
        raise PermissionDeniedError()
    return crud_customer.add_contact(db, customer_id=customer_id, obj_in=body)


@router.post("/{customer_id}/follow-ups", response_model=CustomerFollowUpOut, status_code=201)
def add_follow_up(
    customer_id: int,
    body: CustomerFollowUpCreate,
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_UPDATE)),
    db: Session = Depends(get_db),
):
    customer = crud_customer.get(db, id=customer_id)
    if not customer or customer.is_deleted:
        raise NotFoundError("客户")
    if not check_data_scope(customer, current_user):
        raise PermissionDeniedError()
    return crud_customer.add_follow_up(
        db, customer_id=customer_id, creator_id=current_user.id, obj_in=body
    )


@router.get("/{customer_id}/follow-ups", response_model=list[CustomerFollowUpOut])
def list_follow_ups(
    customer_id: int,
    current_user: User = Depends(require_permissions(PermissionCode.CUSTOMER_READ)),
    db: Session = Depends(get_db),
):
    customer = crud_customer.get(db, id=customer_id)
    if not customer or customer.is_deleted:
        raise NotFoundError("客户")
    if not check_data_scope(customer, current_user):
        raise PermissionDeniedError()
    return crud_customer.get_follow_ups(db, customer_id=customer_id)
